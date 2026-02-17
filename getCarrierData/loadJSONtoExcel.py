#!/usr/bin/env python3
import argparse
import datetime
import json
import logging
import re
from openpyxl import load_workbook
from sys import exit, stdout, exc_info
import traceback

# DataDir = '/Users/jburgess/Library/CloudStorage/Dropbox/CarrierDataCollection/'
# on MY mac, this is an symlink to the above; on the MacMini, this is IT
# DataDir = '/Users/jburgess/Dropbox/CarrierDataCollection/'
# Above are inflexible and not correct on both systems.
# so just use the current directory.
DataDir = "./"
ExcelFile = 'carrier infinity usage stats.xlsm'

# all dates are "yyyy-mm-dd" e.g. "2025-12-13"
datePattern = r'"*\d{4}[/-]\d{1,2}[/-]\d{1,2}"*'
dateRe = re.compile (datePattern)
# all times are "hh:mm:ss" e.g. "12:16:23"
timePattern = r'"*\d{1,2}:\d{1,2}:\d{1,2}"*'
timeRe = re.compile (timePattern)

# fast efficient way to turn numbers as strings into int or float or date or time
def str2num(s):
  try:
      return int(s)
  except:
    try:
      return float(s)
    except:
      try:
        match = dateRe.fullmatch(s)
        if match:
          logging.debug (f"str2num date ({s})")
          return datetime.datetime.strptime(s, "%Y-%m-%d").date()
        else:
          match = timeRe.fullmatch(s)
          if match:
            logging.debug (f"str2num time ({s})")
            return datetime.datetime.strptime(s, "%H:%M:%S").time()
          else:
            return(s)
      except:
        excType, excValue, excTraceback = exc_info()
        lines = traceback.format_exception(excType, excValue, excTraceback)
        logging.warning (f"str2num: triple exception. ({s})\n\tType={excType}, Value={excValue}")
        logging.debug( "".join( lines ))
        return s

def loadJsonToExcel (jsonFile: str, sheet_name: str):
  logging.debug (f"Read Json file {DataDir}{jsonFile}")

  wb = load_workbook(DataDir + ExcelFile)
  logging.debug (f"worksheets in {ExcelFile} are: {wb.sheetnames}")

  ws = wb[sheet_name]
  logging.info (f"sheet {ws.title} has {ws.max_row} rows to start")

  # the list of fields we want to collect from the JSON is in row 1
  field_list = list(next(ws.iter_rows(values_only=True, max_row=1)))
  num_fields = len(field_list)
  logging.debug(f"{num_fields} field names in row 1: {field_list}")

  # now load all the saved data into excel, one row at a time, arranged by the field_list columns
  with open(DataDir + jsonFile, 'r') as jf:
    new_line_count = 0
    for line in jf:
      new_row = [None] * num_fields
      input_dict = json.loads(line)
      new_line_count +=1
      i = -1 # because of the PRE-increment
      for field in field_list:
        i+=1
        if (field == None) or (field == "") or (field[0] == '*'):
          continue
        if field in input_dict:
          new_row[i] = str2num ( input_dict[field] )
        else:
          logging.error(f"line {new_line_count} in the input is missing field {field}")

      logging.debug(f"new row #{new_line_count}: {new_row}")
      ws.append(new_row)

  wb.save(DataDir + ExcelFile)
  logging.info (f"appended {new_line_count} rows")

##########
def main():
  parser = argparse.ArgumentParser(
    description="Load collected JSON into the excel spreadsheet"
  )
  # Example argument; add more as needed
  parser.add_argument( "-d", "--debug", action="store_true", help="Enable debug output" )
  parser.add_argument( "-R", "--RealTime", action="store_true", help="Load JSON RealTime data" )
  parser.add_argument( "-D", "--Daily", action="store_true", help="Load JSON Daily Fields" )
  args = parser.parse_args()

  if args.debug:
    logging.basicConfig(level=logging.DEBUG)
    logging.debug("Debug mode enabled.")
  else:
    logging.basicConfig(level=logging.INFO)
  logging.debug ("Args=[ %s ]" % args)

  if args.RealTime and args.Daily:
    logging.error ("You must specify only ONE of RealTime and Daily")
    exit(1)
  elif args.RealTime:
    loadJsonToExcel ( 'CarrierRealTimeData.json', 'RealTime' )
  elif args.Daily:
    loadJsonToExcel ( 'CarrierDailyData.json', 'Daily' )
  else:
    logging.error ("You must specify either --RealTime or --Daily")
    exit(1)

if __name__ == "__main__":
  main()
